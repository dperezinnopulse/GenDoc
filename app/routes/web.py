from fastapi import APIRouter, Depends, Form, UploadFile, File, Query, Request
from concurrent.futures import ThreadPoolExecutor
import asyncio
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse, Response
from starlette import status
from ..utils.auth import get_session_user, login_user, require_user, set_auth_cookie, clear_auth_cookie
from ..services.template_store import TemplateStore
from ..services.renderer import Renderer
from ..utils.pdf_preview import render_pdf_page_png, get_preview_scale
from jinja2 import Environment, FileSystemLoader, select_autoescape
import io
import json
from pathlib import Path
import re
from pypdf import PdfReader
from openpyxl import load_workbook

router = APIRouter()

BASE_DIR = Path(__file__).resolve().parents[2]

templates_env = Environment(
    loader=FileSystemLoader(str(BASE_DIR / "app" / "templates")),
    autoescape=select_autoescape(["html", "xml"]),
)

store = TemplateStore(base_path="storage/templates")
renderer = Renderer(store)


def _ensure_path(d: dict, path: str):
    cur = d
    parts = [p for p in path.split('.') if p]
    for p in parts[:-1]:
        if p not in cur or not isinstance(cur.get(p), dict):
            cur[p] = {}
        cur = cur[p]
    return cur, parts[-1] if parts else None


def _sample_value_for(field_name: str) -> str:
    low = field_name.lower()
    if "nombre" in low:
        return "Ana"
    if "apellido" in low:
        return "Pérez"
    if "documento" in low or "dni" in low or "numero" in low:
        return "123"
    if "fecha" in low:
        return "2025-09-01"
    if "curso" in low:
        return "Matemáticas"
    if "imagen" in low or "foto" in low or "logo" in low or "imagen" in low:
        # Return URL example for image fields
        return "https://via.placeholder.com/150x100/0066cc/ffffff?text=Logo"
    return "Texto"


def _normalize_mapping(m):
    if isinstance(m, list):
        if len(m) == 1 and isinstance(m[0], dict):
            return m[0]
        # merge list of dicts, last wins
        merged = {}
        for item in m:
            if isinstance(item, dict):
                merged.update(item)
        return merged
    return m or {}


@router.get("/admin", response_class=HTMLResponse)
async def admin_home(user: str | None = Depends(get_session_user)):
    if not user:
        template = templates_env.get_template("login.html")
        return template.render()
    templates = store.list_templates()
    template = templates_env.get_template("dashboard.html")
    return template.render(user=user, templates=templates)

@router.post("/admin/login")
async def admin_login(username: str = Form(...), password: str = Form(...)):
    token = login_user(username, password)
    if token:
        resp = RedirectResponse("/admin", status_code=status.HTTP_302_FOUND)
        set_auth_cookie(resp, token)
        return resp
    template = templates_env.get_template("login.html")
    return HTMLResponse(template.render(error="Credenciales inválidas"), status_code=401)

@router.get("/admin/logout")
async def admin_logout():
    resp = RedirectResponse("/admin", status_code=status.HTTP_302_FOUND)
    clear_auth_cookie(resp)
    return resp

@router.post("/admin/templates/upload")
async def upload_template(
    user: str = Depends(require_user),
    file: UploadFile = File(...),
    name: str = Form(None)
):
    store.save_template(file, name)
    return RedirectResponse(f"/admin", status_code=status.HTTP_302_FOUND)

@router.get("/admin/templates/{template_id}", response_class=HTMLResponse)
async def view_template(template_id: str, user: str = Depends(require_user)):
    template_meta = store.get_template_meta(template_id)
    mapping_raw = template_meta.get("mapping", {})
    mapping = _normalize_mapping(mapping_raw)
    positions: dict = mapping.get("_positions", {}) or {}
    repeat_rows: dict = mapping.get("_repeat_rows", {}) or {}

    # Build suggested data
    sample: dict = {}
    # Arrays first
    for arr_path, _def in repeat_rows.items():
        items_fields = [k[len(arr_path)+1:] for k in positions.keys() if k.startswith(arr_path + ".")]
        def build_item():
            item: dict = {}
            for subkey in items_fields:
                parent, leaf = _ensure_path(item, subkey)
                if leaf:
                    parent[leaf] = _sample_value_for(leaf)
            if not item:
                item["valor"] = "Texto"
            return item
        sample[arr_path] = [build_item(), build_item()]
    # Fixed fields (not under arrays)
    for key in positions.keys():
        if any(key.startswith(p + ".") for p in repeat_rows.keys()):
            continue
        parent, leaf = _ensure_path(sample, key)
        if leaf:
            parent[leaf] = _sample_value_for(leaf)
    # Also include plain mapping keys (non-reserved)
    for k, v in mapping.items():
        if k.startswith("_"):
            continue
        if k not in sample:
            sample[k] = _sample_value_for(k)
    
    # Include image fields from _images mapping
    images_cfg = mapping.get("_images", {}) or {}
    for img_key in images_cfg.keys():
        if img_key not in sample:
            sample[img_key] = _sample_value_for(img_key)

    # Fallbacks if still empty: inspect template to propose fields
    if not sample:
        kind = template_meta.get("kind")
        tpl_path = store.get_template_file(template_id)
        if kind == "pdf":
            try:
                reader = PdfReader(tpl_path)
                fields = None
                if hasattr(reader, "get_fields"):
                    fields = reader.get_fields() or {}
                if fields:
                    for fname in fields.keys():
                        parent, leaf = _ensure_path(sample, fname)
                        if leaf:
                            parent[leaf] = _sample_value_for(leaf)
            except Exception:
                pass
        elif kind == "xlsx":
            try:
                wb = load_workbook(tpl_path)
                placeholder_re = re.compile(r"\{\{\s*([^}]+?)\s*\}\}")
                found = set()
                for ws in wb.worksheets:
                    for row in ws.iter_rows():
                        for cell in row:
                            if isinstance(cell.value, str):
                                for m in placeholder_re.findall(cell.value):
                                    found.add(m)
                for key in sorted(found):
                    parent, leaf = _ensure_path(sample, key)
                    if leaf:
                        parent[leaf] = _sample_value_for(leaf)
            except Exception:
                pass

    suggested_data = json.dumps(sample if sample else {"nombre": "Ana"}, ensure_ascii=False, indent=2)

    # override mapping shown to be normalized for the UI
    tm = dict(template_meta)
    tm["mapping"] = mapping

    template = templates_env.get_template("template_detail.html")
    return template.render(user=user, template=tm, suggested_data=suggested_data)

@router.post("/admin/templates/{template_id}/delete")
async def delete_template(template_id: str, user: str = Depends(require_user)):
    store.delete_template(template_id)
    return RedirectResponse("/admin", status_code=status.HTTP_302_FOUND)

@router.post("/admin/templates/{template_id}/save_mapping")
async def save_mapping(
    template_id: str,
    mapping_json: str = Form("{}"),
    repeat_json: str = Form("{}"),
    schema_json: str = Form("{}"),
    user: str = Depends(require_user),
):
    try:
        mapping_loaded = json.loads(mapping_json) if mapping_json else {}
        mapping = _normalize_mapping(mapping_loaded)
        repeat_sections = json.loads(repeat_json) if repeat_json else {}
        schema = json.loads(schema_json) if schema_json else {}
    except json.JSONDecodeError as ex:
        tmpl = templates_env.get_template("template_detail.html")
        template_meta = store.get_template_meta(template_id)
        return HTMLResponse(tmpl.render(user=user, template=template_meta, error=f"JSON inválido: {ex}"), status_code=400)
    store.save_mapping(template_id, mapping, repeat_sections, schema)
    return RedirectResponse(f"/admin/templates/{template_id}", status_code=status.HTTP_302_FOUND)

@router.post("/admin/templates/{template_id}/test_render")
async def test_render(template_id: str, data_json: str = Form("{}"), format: str = Form("pdf"), user: str = Depends(require_user)):
    try:
        data = json.loads(data_json) if data_json else {}
        
        if format == "image":
            # Generar imagen PNG
            pdf_bytes = renderer.render_to_pdf(template_id, data)
            # Usar un executor para la conversión síncrona
            loop = asyncio.get_event_loop()
            with ThreadPoolExecutor() as executor:
                image_bytes = await loop.run_in_executor(
                    executor, 
                    renderer._convert_pdf_to_image_sync, 
                    pdf_bytes
                )
            return StreamingResponse(
                io.BytesIO(image_bytes), 
                media_type="image/png", 
                headers={"Content-Disposition": f"inline; filename=debug-{template_id}.png"}
            )
        else:
            # Generar PDF (por defecto)
            pdf_bytes = renderer.render_to_pdf(template_id, data)
            return StreamingResponse(
                io.BytesIO(pdf_bytes), 
                media_type="application/pdf", 
                headers={"Content-Disposition": f"inline; filename=debug-{template_id}.pdf"}
            )
    except Exception as ex:
        tmpl = templates_env.get_template("template_detail.html")
        template_meta = store.get_template_meta(template_id)
        return HTMLResponse(tmpl.render(user=user, template=template_meta, error=str(ex)), status_code=400)

@router.get("/admin/templates/{template_id}/test_render")
async def test_render_get(template_id: str, format: str = Query("pdf"), user: str = Depends(require_user)):
    # Compute suggested data and return a debug PDF directly
    template_meta = store.get_template_meta(template_id)
    mapping = _normalize_mapping(template_meta.get("mapping", {}))
    positions: dict = mapping.get("_positions", {}) or {}
    repeat_rows: dict = mapping.get("_repeat_rows", {}) or {}

    # Build suggested data (same as detail view)
    sample: dict = {}
    for arr_path, _def in repeat_rows.items():
        items_fields = [k[len(arr_path)+1:] for k in positions.keys() if k.startswith(arr_path + ".")]
        def build_item():
            item: dict = {}
            for subkey in items_fields:
                parent, leaf = _ensure_path(item, subkey)
                if leaf:
                    parent[leaf] = _sample_value_for(leaf)
            if not item:
                item["valor"] = "Texto"
            return item
        sample[arr_path] = [build_item(), build_item()]
    for key in positions.keys():
        if any(key.startswith(p + ".") for p in repeat_rows.keys()):
            continue
        parent, leaf = _ensure_path(sample, key)
        if leaf:
            parent[leaf] = _sample_value_for(leaf)
    for k, v in mapping.items():
        if k.startswith("_"):
            continue
        if k not in sample:
            parent, leaf = _ensure_path(sample, k)
            if leaf:
                parent[leaf] = _sample_value_for(leaf)
            else:
                sample[k] = _sample_value_for(k)

    if not sample:
        kind = template_meta.get("kind")
        tpl_path = store.get_template_file(template_id)
        if kind == "pdf":
            try:
                reader = PdfReader(tpl_path)
                fields = None
                if hasattr(reader, "get_fields"):
                    fields = reader.get_fields() or {}
                if fields:
                    for fname in fields.keys():
                        parent, leaf = _ensure_path(sample, fname)
                        if leaf:
                            parent[leaf] = _sample_value_for(leaf)
            except Exception:
                pass
        elif kind == "xlsx":
            try:
                wb = load_workbook(tpl_path)
                placeholder_re = re.compile(r"\{\{\s*([^}]+?)\s*\}\}")
                found = set()
                for ws in wb.worksheets:
                    for row in ws.iter_rows():
                        for cell in row:
                            if isinstance(cell.value, str):
                                for m in placeholder_re.findall(cell.value):
                                    found.add(m)
                for key in sorted(found):
                    parent, leaf = _ensure_path(sample, key)
                    if leaf:
                        parent[leaf] = _sample_value_for(leaf)
            except Exception:
                pass
    if not sample:
        sample = {"nombre": "Ana"}

    if format == "image":
        # Generar imagen PNG
        pdf_bytes = renderer.render_to_pdf(template_id, sample)
        # Usar un executor para la conversión síncrona
        loop = asyncio.get_event_loop()
        with ThreadPoolExecutor() as executor:
            image_bytes = await loop.run_in_executor(
                executor, 
                renderer._convert_pdf_to_image_sync, 
                pdf_bytes
            )
        return StreamingResponse(
            io.BytesIO(image_bytes), 
            media_type="image/png", 
            headers={"Content-Disposition": f"inline; filename=debug-get-{template_id}.png"}
        )
    else:
        # Generar PDF (por defecto)
        pdf_bytes = renderer.render_to_pdf(template_id, sample)
        return StreamingResponse(
            io.BytesIO(pdf_bytes), 
            media_type="application/pdf", 
            headers={"Content-Disposition": f"inline; filename=debug-get-{template_id}.pdf"}
        )

# Overlay editor
@router.get("/admin/templates/{template_id}/overlay", response_class=HTMLResponse)
async def overlay_editor(template_id: str, user: str = Depends(require_user)):
    template_meta = store.get_template_meta(template_id)
    if template_meta.get("kind") != "pdf":
        return RedirectResponse(f"/admin/templates/{template_id}")
    template = templates_env.get_template("overlay_editor.html")
    mp = _normalize_mapping(template_meta.get("mapping", {}))
    positions = mp.get("_positions", {})
    repeat_rows = mp.get("_repeat_rows", {})
    styles = mp.get("_styles", {})
    default_style = mp.get("_default_style", {"font":"Helvetica","size":10,"color":"#111111"})
    header_positions = mp.get("_header_positions", {})
    footer_positions = mp.get("_footer_positions", {})
    offset = mp.get("_offset", {"x": 0, "y": 0})
    preview_scale = mp.get("_preview_scale", get_preview_scale())
    images = mp.get("_images", {})
    image_previews = mp.get("_image_previews", {})
    # Fallback to meta-level if mapping doesn't have them
    if not images:
        images = template_meta.get("_images", {})
    if not image_previews:
        image_previews = template_meta.get("_image_previews", {})
    return template.render(
        user=user,
        template=template_meta,
        positions=json.dumps(positions),
        repeat_rows=json.dumps(repeat_rows),
        styles=json.dumps(styles),
        default_style=json.dumps(default_style),
        header_positions=json.dumps(header_positions),
        footer_positions=json.dumps(footer_positions),
        preview_scale=preview_scale,
        offset_x=offset.get("x", 0),
        offset_y=offset.get("y", 0),
        images=json.dumps(images),
        image_previews=json.dumps(image_previews),
    )

@router.get("/admin/templates/{template_id}/overlay/page.png")
async def overlay_editor_page_png(template_id: str, page: int = Query(1, ge=1), user: str = Depends(require_user)):
    path = store.get_template_file(template_id)
    img_bytes = render_pdf_page_png(path, page-1)
    return Response(content=img_bytes, media_type="image/png")

@router.post("/admin/templates/{template_id}/overlay/save")
async def overlay_editor_save(template_id: str, positions_json: str = Form("{}"), repeat_rows_json: str = Form("{}"), header_json: str = Form(None), footer_json: str = Form(None), styles_json: str = Form(None), default_style_json: str = Form(None), offset_json: str = Form(None), preview_scale: float = Form(None), images_json: str = Form("{}"), image_previews_json: str = Form("{}"), user: str = Depends(require_user)):
    meta = store.get_template_meta(template_id)
    try:
        positions = json.loads(positions_json) if positions_json else {}
    except json.JSONDecodeError:
        positions = {}
    try:
        repeat_rows = json.loads(repeat_rows_json) if repeat_rows_json else {}
    except json.JSONDecodeError:
        repeat_rows = {}
    header_positions = {}
    footer_positions = {}
    styles = {}
    default_style = None
    try:
        if header_json:
            header_positions = json.loads(header_json)
    except json.JSONDecodeError:
        header_positions = {}
    try:
        if footer_json:
            footer_positions = json.loads(footer_json)
    except json.JSONDecodeError:
        footer_positions = {}
    try:
        if styles_json:
            styles = json.loads(styles_json)
    except json.JSONDecodeError:
        styles = {}
    try:
        if default_style_json:
            default_style = json.loads(default_style_json)
    except json.JSONDecodeError:
        default_style = None
    try:
        offset = json.loads(offset_json) if offset_json else None
    except json.JSONDecodeError:
        offset = None
    try:
        images = json.loads(images_json) if images_json else {}
    except json.JSONDecodeError:
        images = {}
    try:
        image_previews = json.loads(image_previews_json) if image_previews_json else {}
    except json.JSONDecodeError:
        image_previews = {}

    mapping = _normalize_mapping(meta.get("mapping", {}))
    mapping["_positions"] = positions
    mapping["_repeat_rows"] = repeat_rows
    if header_positions:
        mapping["_header_positions"] = header_positions
    if footer_positions:
        mapping["_footer_positions"] = footer_positions
    if styles is not None:
        mapping["_styles"] = styles
    if default_style is not None:
        mapping["_default_style"] = default_style
    if offset is not None:
        mapping["_offset"] = {"x": float(offset.get("x", 0)), "y": float(offset.get("y", 0))}
    if preview_scale:
        mapping["_preview_scale"] = float(preview_scale)
    if images is not None:
        mapping["_images"] = images
    if image_previews is not None:
        mapping["_image_previews"] = image_previews

    store.save_mapping(template_id, mapping, meta.get("repeat_sections", {}), meta.get("schema", {}), images=images, image_previews=image_previews)
    return RedirectResponse(f"/admin/templates/{template_id}/overlay", status_code=302)

@router.get("/admin/templates/{template_id}/markers", response_class=HTMLResponse)
async def markers_editor(template_id: str, user: str = Depends(require_user)):
    template_meta = store.get_template_meta(template_id)
    mapping = _normalize_mapping(template_meta.get("mapping", {}))
    positions = mapping.get("_positions", {}) or {}
    header_positions = mapping.get("_header_positions", {}) or {}
    footer_positions = mapping.get("_footer_positions", {}) or {}
    images = mapping.get("_images", {}) or {}
    signatures = mapping.get("_signatures", {}) or {}
    template = templates_env.get_template("markers_editor.html")
    return template.render(user=user, template=template_meta, positions=positions, header_positions=header_positions, footer_positions=footer_positions, images=images, signatures=signatures)

@router.post("/admin/templates/{template_id}/markers")
async def save_markers(template_id: str, request: Request, user: str = Depends(require_user)):
    form = await request.form()
    mapping = _normalize_mapping(store.get_template_meta(template_id).get("mapping", {}))
    
    # Collect fixed markers
    names = form.getlist("name[]")
    xs = form.getlist("x[]")
    ys = form.getlist("y[]")
    types = form.getlist("type[]")
    new_positions: dict[str, list[float]] = {}
    for i in range(len(names)):
        k = (names[i] or "").strip()
        if not k:
            continue
        try:
            x = float(xs[i])
            y = float(ys[i])
        except Exception:
            continue
        new_positions[k] = [x, y]
    
    # Collect image markers
    img_names = form.getlist("img_name[]")
    img_xs = form.getlist("img_x[]")
    img_ys = form.getlist("img_y[]")
    img_widths = form.getlist("img_width[]")
    img_heights = form.getlist("img_height[]")
    new_images: dict[str, dict] = {}
    for i in range(len(img_names)):
        k = (img_names[i] or "").strip()
        if not k:
            continue
        try:
            x = float(img_xs[i])
            y = float(img_ys[i])
            width = float(img_widths[i])
            height = float(img_heights[i])
        except Exception:
            continue
        new_images[k] = {"x": x, "y": y, "width": width, "height": height}

    # Collect signature markers
    sig_names = form.getlist("sig_name[]")
    sig_xs = form.getlist("sig_x[]")
    sig_ys = form.getlist("sig_y[]")
    sig_widths = form.getlist("sig_width[]")
    sig_heights = form.getlist("sig_height[]")
    new_signatures: dict[str, dict] = {}
    for i in range(len(sig_names)):
        k = (sig_names[i] or "").strip()
        if not k:
            continue
        try:
            x = float(sig_xs[i])
            y = float(sig_ys[i])
            width = float(sig_widths[i]) if sig_widths[i] else 200.0
            height = float(sig_heights[i]) if sig_heights[i] else 300.0
        except Exception:
            continue
        new_signatures[k] = {"x": x, "y": y, "width": width, "height": height}
    
    # Header markers
    h_names = form.getlist("hname[]")
    h_xs = form.getlist("hx[]")
    h_ys = form.getlist("hy[]")
    new_header: dict[str, list[float]] = {}
    for i in range(len(h_names)):
        k = (h_names[i] or "").strip()
        if not k:
            continue
        try:
            x = float(h_xs[i])
            y = float(h_ys[i])
        except Exception:
            continue
        new_header[k] = [x, y]
    
    # Footer markers
    f_names = form.getlist("fname[]")
    f_xs = form.getlist("fx[]")
    f_ys = form.getlist("fy[]")
    new_footer: dict[str, list[float]] = {}
    for i in range(len(f_names)):
        k = (f_names[i] or "").strip()
        if not k:
            continue
        try:
            x = float(f_xs[i])
            y = float(f_ys[i])
        except Exception:
            continue
        new_footer[k] = [x, y]

    mapping["_positions"] = new_positions
    mapping["_header_positions"] = new_header
    mapping["_footer_positions"] = new_footer
    mapping["_images"] = new_images
    mapping["_signatures"] = new_signatures

    meta = store.get_template_meta(template_id)
    store.save_mapping(template_id, mapping, meta.get("repeat_sections", {}), meta.get("schema", {}))
    return RedirectResponse(f"/admin/templates/{template_id}/markers", status_code=302)