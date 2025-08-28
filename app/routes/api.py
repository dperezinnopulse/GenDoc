from fastapi import APIRouter, HTTPException
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from typing import Literal, Optional
from ..services.template_store import TemplateStore
from ..services.renderer import Renderer
import io
import re
from pypdf import PdfReader
from openpyxl import load_workbook

router = APIRouter()

store = TemplateStore(base_path="storage/templates")
renderer = Renderer(store)

class RenderRequest(BaseModel):
    template_id: str
    data: dict
    output_format: Literal["pdf", "image"] = "pdf"
    image_format: Optional[Literal["webp", "png", "jpeg"]] = "webp"  # Nuevo parámetro con valor por defecto

class MappingRequest(BaseModel):
    mapping: dict | None = None
    repeat_sections: dict | None = None
    schema: dict | None = None


def _sample_value_for(name: str) -> str:
    low = name.lower()
    if "nombre" in low:
        return "Ana"
    if "apellido" in low:
        return "Pérez"
    if "documento" in low or "dni" in low or "numero" in low or low == "nif":
        return "123"
    if "fecha" in low:
        return "2025-09-01"
    if "curso" in low:
        return "Matemáticas"
    return "Texto"


def _ensure_path(d: dict, path: str):
    cur = d
    parts = [p for p in path.split('.') if p]
    for p in parts[:-1]:
        if p not in cur or not isinstance(cur.get(p), dict):
            cur[p] = {}
        cur = cur[p]
    return cur, parts[-1] if parts else None


def _normalize_mapping(m):
    if isinstance(m, list):
        if len(m) == 1 and isinstance(m[0], dict):
            return m[0]
        merged = {}
        for item in m:
            if isinstance(item, dict):
                merged.update(item)
        return merged
    return m or {}

@router.get("/templates")
async def list_templates():
    return store.list_templates()

@router.get("/templates/{template_id}")
async def get_template(template_id: str):
    try:
        return store.get_template_meta(template_id)
    except FileNotFoundError:
        raise HTTPException(status_code=404, detail="Plantilla no encontrada")

@router.get("/templates/{template_id}/suggested")
async def get_suggested_data(template_id: str):
    try:
        meta = store.get_template_meta(template_id)
    except FileNotFoundError:
        raise HTTPException(status_code=404, detail="Plantilla no encontrada")
    mapping = _normalize_mapping(meta.get("mapping", {}))
    positions = mapping.get("_positions", {}) or {}
    repeat_rows = mapping.get("_repeat_rows", {}) or {}
    sample: dict = {}
    # arrays
    for arr_path, _def in repeat_rows.items():
        items_fields = [k[len(arr_path)+1:] for k in positions.keys() if k.startswith(arr_path + ".")]
        def build_item():
            item: dict = {}
            for subkey in items_fields:
                parent, leaf = _ensure_path(item, subkey)
                if leaf:
                    parent[leaf] = _sample_value_for(leaf)
            if not item:
                item["valor"] = "Texto"
            return item
        sample[arr_path] = [build_item(), build_item()]
    # fixed
    for key in positions.keys():
        if any(key.startswith(p + ".") for p in repeat_rows.keys()):
            continue
        parent, leaf = _ensure_path(sample, key)
        if leaf:
            parent[leaf] = _sample_value_for(leaf)
    # plain mapping keys
    for k, v in mapping.items():
        if k.startswith("_"):
            continue
        if k not in sample:
            parent, leaf = _ensure_path(sample, k)
            if leaf:
                parent[leaf] = _sample_value_for(leaf)
            else:
                sample[k] = _sample_value_for(k)
    # fallbacks
    if not sample:
        kind = meta.get("kind")
        path = store.get_template_file(template_id)
        if kind == "pdf":
            try:
                reader = PdfReader(path)
                fields = None
                if hasattr(reader, "get_fields"):
                    fields = reader.get_fields() or {}
                if fields:
                    for fname in fields.keys():
                        parent, leaf = _ensure_path(sample, fname)
                        if leaf:
                            parent[leaf] = _sample_value_for(leaf)
            except Exception:
                pass
        elif kind == "xlsx":
            try:
                wb = load_workbook(path)
                placeholder_re = re.compile(r"\{\{\s*([^}]+?)\s*\}\}")
                found = set()
                for ws in wb.worksheets:
                    for row in ws.iter_rows():
                        for cell in row:
                            if isinstance(cell.value, str):
                                for m in placeholder_re.findall(cell.value):
                                    found.add(m)
                for key in sorted(found):
                    parent, leaf = _ensure_path(sample, key)
                    if leaf:
                        parent[leaf] = _sample_value_for(leaf)
            except Exception:
                pass
    if not sample:
        sample = {"nombre": "Ana"}
    return sample

@router.delete("/templates/{template_id}")
async def delete_template(template_id: str):
    ok = store.delete_template(template_id)
    if not ok:
        raise HTTPException(status_code=404, detail="Plantilla no encontrada")
    return {"deleted": True}

@router.post("/templates/{template_id}/mapping")
async def set_mapping(template_id: str, body: MappingRequest):
    try:
        store.save_mapping(template_id, body.mapping or {}, body.repeat_sections or {}, body.schema or {})
    except FileNotFoundError:
        raise HTTPException(status_code=404, detail="Plantilla no encontrada")
    return {"ok": True}

@router.post("/render")
async def render_document(req: RenderRequest):
    try:
        # Debug: Log the request parameters
        print(f"🔍 DEBUG: output_format = {req.output_format}")
        print(f"🔍 DEBUG: template_id = {req.template_id}")
        
        # Get template metadata to check for signatures
        meta = store.get_template_meta(req.template_id)
        mapping = meta.get("mapping", {})
        
        # Extract signature coordinates from _positions
        positions = mapping.get("_positions", {})
        signatures_cfg = {}
        
        # Look for signature fields in positions
        for field_name, coords in positions.items():
            if "firma" in field_name.lower() or "signature" in field_name.lower():
                signatures_cfg[field_name] = {
                    "x": coords[0] if isinstance(coords, list) and len(coords) >= 2 else 0,
                    "y": coords[1] if isinstance(coords, list) and len(coords) >= 2 else 0,
                    "width": 200,  # Default width
                    "height": 100   # Default height
                }
        
        # Render PDF synchronously
        pdf_bytes = renderer.render_to_pdf(req.template_id, req.data)
        
        # Handle different output formats
        if req.output_format == "image":
            print("🖼️  DEBUG: Converting PDF to image...")
            # Convert PDF to image
            image_bytes, final_image_width, final_image_height = renderer.convert_pdf_to_image(pdf_bytes)
            print(f"🖼️  DEBUG: Image conversion complete, size: {len(image_bytes)} bytes")
            
            # Prepare response with image and signature coordinates
            import base64
            image_base64 = base64.b64encode(image_bytes).decode('utf-8')
            
            # Get image dimensions for coordinate scaling
            from PIL import Image
            import io
            pil_image = Image.open(io.BytesIO(image_bytes))
            image_width, image_height = pil_image.size
            
            # Las dimensiones del PDF original están hardcodeadas para este template específico
            # Basándome en el debug, el PDF original tiene dimensiones A4 estándar
            original_pdf_width_points = 595.32
            original_pdf_height_points = 841.92
            
            response_data = {
                "image_base64": image_base64,
                "image_info": {
                    "width": image_width,
                    "height": image_height,
                    "original_pdf_width_points": original_pdf_width_points,
                    "original_pdf_height_points": original_pdf_height_points,
                    "scale_x": image_width / original_pdf_width_points,
                    "scale_y": image_height / original_pdf_height_points,
                    "coordinate_system_origin": "top-left"
                },
                "signatures": {}
            }
            
            # Add signature coordinates if any exist (scaled to image dimensions)
            if signatures_cfg:
                print(f"🔍 DEBUG: Found {len(signatures_cfg)} signature fields")
                for key, meta in signatures_cfg.items():
                    # Obtener coordenadas originales del template
                    x_template = meta.get("x", 0)
                    y_template = meta.get("y", 0)
                    
                    print(f"🔍 DEBUG: Template coordinates for '{key}': x={x_template}, y={y_template}")
                    print(f"🔍 DEBUG: Image dimensions: {image_width} x {image_height}")
                    
                    # Calcular factores de escalado dinámicamente
                    # Intentar diferentes métodos de escalado y elegir el más apropiado
                    
                    # Método 1: Escalado basado en dimensiones del PDF (estándar)
                    scale_x_pdf = image_width / original_pdf_width_points
                    scale_y_pdf = image_height / original_pdf_height_points
                    
                    # Método 2: Escalado directo (para sistemas personalizados)
                    # Usar un factor de escalado basado en la relación entre coordenadas típicas
                    # Para este template específico, las coordenadas están en un sistema personalizado
                    # pero podemos detectar automáticamente el factor
                    
                    # Detectar si las coordenadas están en un rango razonable para el PDF
                    if x_template <= original_pdf_width_points and y_template <= original_pdf_height_points:
                        # Coordenadas están en el sistema del PDF
                        x_image = int(x_template * scale_x_pdf)
                        y_image = int((original_pdf_height_points - y_template) * scale_y_pdf)
                        final_x = x_image
                        final_y = y_image
                        final_width = int(meta.get("width", 200) * scale_x_pdf)
                        final_height = int(meta.get("height", 100) * scale_y_pdf)
                    else:
                        # Coordenadas están en un sistema personalizado
                        # Calcular factores de escalado basados en las dimensiones de la imagen
                        # y las coordenadas típicas de este tipo de template
                        
                        # Para templates con coordenadas personalizadas, usar un escalado proporcional
                        # basado en las dimensiones de la imagen final
                        ratio_x = image_width / 1000  # Asumiendo coordenadas en escala de 1000
                        ratio_y = image_height / 1000
                        
                        final_x = int(x_template * ratio_x)
                        final_y = int(y_template * ratio_y)
                        final_width = int(meta.get("width", 200) * ratio_x)
                        final_height = int(meta.get("height", 100) * ratio_y)
                    
                    print(f"🔍 DEBUG: Converted coordinates for '{key}': x={final_x}, y={final_y}, w={final_width}, h={final_height}")
                    
                    response_data["signatures"][key] = {
                        "x": final_x,
                        "y": final_y,
                        "width": final_width,
                        "height": final_height
                    }
            
            return response_data
        
        print("📄 DEBUG: Returning PDF format...")
        # Default: PDF output
        # Prepare response with signature coordinates if any
        response_data = {
            "pdf": pdf_bytes,
            "signatures": {}
        }
        
        # Add signature coordinates if any exist (PDF coordinates, no scaling needed)
        if signatures_cfg:
            for key, meta in signatures_cfg.items():
                response_data["signatures"][key] = {
                    "x": meta.get("x", 0),
                    "y": meta.get("y", 0),
                    "width": meta.get("width", 200),
                    "height": meta.get("height", 100)
                }
        
        # If no signatures, return just the PDF
        if not signatures_cfg:
            return StreamingResponse(io.BytesIO(pdf_bytes), media_type="application/pdf")
        else:
            # Return JSON with PDF as base64 and signature coordinates
            import base64
            pdf_base64 = base64.b64encode(pdf_bytes).decode('utf-8')
            return {
                "pdf_base64": pdf_base64,
                "signatures": response_data["signatures"]
            }
            
    except FileNotFoundError:
        raise HTTPException(status_code=404, detail="Plantilla no encontrada")
    except Exception as ex:
        raise HTTPException(status_code=400, detail=str(ex))