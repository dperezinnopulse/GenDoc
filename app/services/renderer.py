import os
import tempfile
import math
from typing import Dict, Any, List
from .template_store import TemplateStore
from ..utils.soffice import convert_to_pdf
from ..utils.validation import validate_payload
from ..utils.pdf_preview import get_preview_scale
from docxtpl import DocxTemplate
from openpyxl import load_workbook
from reportlab.pdfgen import canvas
from reportlab.lib.colors import HexColor
from reportlab.lib.utils import ImageReader
from pypdf import PdfReader, PdfWriter
import io
import asyncio
from concurrent.futures import ThreadPoolExecutor


class Renderer:
    def __init__(self, store: TemplateStore):
        self.store = store

    def _render_to_pdf_sync(self, template_id: str, data: Dict[str, Any]) -> bytes:
        """Versión síncrona del renderizado para usar cuando ya hay un event loop corriendo."""
        template_meta = self.store.get_template_meta(template_id)
        if not template_meta:
            raise ValueError(f"Template {template_id} not found")
        
        # Aplicar mapping
        mapping = template_meta.get("mapping", {})
        context = self._apply_mapping(data, mapping)
        
        # Renderizar según el tipo
        kind = template_meta.get("kind")
        tpl_path = self.store.get_template_file(template_id)
        
        if kind == "docx":
            return self._render_docx_to_pdf_sync(tpl_path, context)
        elif kind == "xlsx":
            return self._render_xlsx_to_pdf(tpl_path, context)
        elif kind == "pdf":
            if mapping.get("_positions"):
                return self._render_pdf_overlay(tpl_path, context, data, mapping)
            else:
                return self._render_pdf_acroform(tpl_path, context)
        else:
            raise ValueError(f"Unsupported template kind: {kind}")

    def _render_docx_to_pdf_sync(self, tpl_path: str, context: Dict[str, Any]) -> bytes:
        """Versión síncrona de la conversión DOCX a PDF."""
        with tempfile.TemporaryDirectory() as td:
            out_docx = os.path.join(td, "out.docx")
            out_pdf = os.path.join(td, "out.pdf")
            doc = DocxTemplate(tpl_path)
            doc.render(context)
            doc.save(out_docx)
            
            # Usar conversión síncrona
            from ..utils.soffice import convert_to_pdf
            success = convert_to_pdf(out_docx, out_pdf)
            
            if not success:
                raise RuntimeError("Error en conversión a PDF")
                
            with open(out_pdf, "rb") as f:
                return f.read()

    async def render_to_pdf_async(self, template_id: str, data: Dict[str, Any]) -> bytes:
        meta = self.store.get_template_meta(template_id)
        kind = meta["kind"]
        tpl_path = self.store.get_template_file(template_id)
        mapping = meta.get("mapping", {})
        repeat_sections = meta.get("repeat_sections", {})
        schema = meta.get("schema")

        validate_payload(data, schema)

        context = self._apply_mapping(data, mapping)
        if kind == "docx":
            return await self._render_docx_to_pdf_async(tpl_path, context)
        if kind == "xlsx":
            return self._render_xlsx_to_pdf(tpl_path, context)  # TODO: Hacer asíncrono
        if kind == "pdf":
            # Prefer overlay if positions mapping exists; otherwise try AcroForm then fallback
            if mapping.get("_positions"):
                return self._render_pdf_overlay(tpl_path, context, original_data=data, mapping=mapping)
            try:
                return self._render_pdf_acroform(tpl_path, context)
            except Exception:
                return self._render_pdf_overlay(tpl_path, context, original_data=data, mapping=mapping)
        raise ValueError("Tipo de plantilla no soportado")

    def render_to_pdf(self, template_id: str, data: Dict[str, Any]) -> bytes:
        # Versión síncrona para compatibilidad
        import asyncio
        try:
            loop = asyncio.get_event_loop()
            if loop.is_running():
                # Si el loop ya está corriendo, usar el método síncrono directamente
                return self._render_to_pdf_sync(template_id, data)
            else:
                return loop.run_until_complete(self.render_to_pdf_async(template_id, data))
        except RuntimeError:
            # Si no hay loop, crear uno nuevo
            return asyncio.run(self.render_to_pdf_async(template_id, data))

    def _apply_mapping(self, data: Dict[str, Any], mapping: Dict[str, Any]) -> Dict[str, Any]:
        if not mapping:
            return data
        def get_path(d: Dict[str, Any], path: str):
            cur: Any = d
            for part in path.split('.'):
                if isinstance(cur, dict):
                    cur = cur.get(part)
                else:
                    cur = None
                if cur is None:
                    break
            return cur
        out: Dict[str, Any] = {}
        for k, src in mapping.items():
            if isinstance(src, str):
                out[k] = get_path(data, src)
        out.update({k: v for k, v in mapping.items() if not isinstance(v, str)})
        out.update(data)
        return out

    async def _render_docx_to_pdf_async(self, tpl_path: str, context: Dict[str, Any]) -> bytes:
        with tempfile.TemporaryDirectory() as td:
            out_docx = os.path.join(td, "out.docx")
            out_pdf = os.path.join(td, "out.pdf")
            doc = DocxTemplate(tpl_path)
            doc.render(context)
            doc.save(out_docx)
            
            # Usar pool de conversión asíncrono
            from ..utils.soffice_pool import get_soffice_pool
            soffice_pool = get_soffice_pool()
            success = await soffice_pool.convert_to_pdf_async(out_docx, out_pdf)
            
            if not success:
                raise RuntimeError("Error en conversión a PDF")
                
            with open(out_pdf, "rb") as f:
                return f.read()
    
    def _render_docx_to_pdf(self, tpl_path: str, context: Dict[str, Any]) -> bytes:
        # Versión síncrona para compatibilidad
        import asyncio
        try:
            loop = asyncio.get_event_loop()
            return loop.run_until_complete(self._render_docx_to_pdf_async(tpl_path, context))
        except RuntimeError:
            # Si no hay loop, crear uno nuevo
            return asyncio.run(self._render_docx_to_pdf_async(tpl_path, context))

    def _render_xlsx_to_pdf(self, tpl_path: str, context: Dict[str, Any]) -> bytes:
        with tempfile.TemporaryDirectory() as td:
            out_xlsx = os.path.join(td, "out.xlsx")
            out_pdf = os.path.join(td, "out.pdf")
            wb = load_workbook(tpl_path)
            for ws in wb.worksheets:
                for row in ws.iter_rows():
                    for cell in row:
                        if isinstance(cell.value, str):
                            new_val = cell.value
                            for k, v in context.items():
                                placeholder = f"{{{{{k}}}}}"
                                if placeholder in new_val:
                                    new_val = new_val.replace(placeholder, str(v) if v is not None else "")
                            cell.value = new_val
            wb.save(out_xlsx)
            convert_to_pdf(out_xlsx, out_pdf)
            with open(out_pdf, "rb") as f:
                return f.read()

    def _render_pdf_acroform(self, tpl_path: str, context: Dict[str, Any]) -> bytes:
        with tempfile.TemporaryDirectory() as td:
            reader = PdfReader(tpl_path)
            writer = PdfWriter()
            for page in reader.pages:
                writer.add_page(page)
            fields = {}
            for key, val in context.items():
                if key.startswith("_"):
                    continue
                fields[key] = "" if val is None else str(val)
            writer.update_page_form_field_values(writer.pages[0], fields)
            out_path = os.path.join(td, "out.pdf")
            with open(out_path, "wb") as f:
                writer.write(f)
            with open(out_path, "rb") as f:
                return f.read()

    def _render_pdf_overlay(self, tpl_path: str, context: Dict[str, Any], original_data: Dict[str, Any], mapping: Dict[str, Any]) -> bytes:
        positions = mapping.get("_positions", {})
        repeat_rows_cfg = mapping.get("_repeat_rows", {})
        header_positions = mapping.get("_header_positions", {})
        footer_positions = mapping.get("_footer_positions", {})
        styles = mapping.get("_styles", {})
        default_style = mapping.get("_default_style", {"font": "Helvetica", "size": 10, "color": "#111111"})

        preview_scale = float(mapping.get("_preview_scale", get_preview_scale()))
        offset_conf = mapping.get("_offset", {"x": 0, "y": 0})
        offset_x = float(offset_conf.get("x", 0) or 0)
        offset_y = float(offset_conf.get("y", 0) or 0)

        def apply_style_for_key(key: str):
            style = styles.get(key, default_style)
            font = style.get("font", default_style.get("font", "Helvetica"))
            size = float(style.get("size", default_style.get("size", 10)))
            color = style.get("color", default_style.get("color", "#111111"))
            try:
                c.setFont(font, size)
            except Exception:
                c.setFont("Helvetica", size)
            try:
                c.setFillColor(HexColor(color))
            except Exception:
                c.setFillColor(HexColor("#111111"))

        def draw_text(key: str, x: float, y: float, value: Any):
            apply_style_for_key(key)
            c.drawString(float(x) + offset_x, float(y) + offset_y, "" if value is None else str(value))

        def draw_cross(x: float, y: float, size: float = 4.0):
            cx = float(x) + offset_x
            cy = float(y) + offset_y
            c.setLineWidth(0.5)
            c.line(cx - size, cy, cx + size, cy)
            c.line(cx, cy - size, cx, cy + size)

        def get_path(d: Dict[str, Any], path: str):
            cur: Any = d
            for part in path.split('.'):
                if isinstance(cur, dict):
                    cur = cur.get(part)
                else:
                    cur = None
                if cur is None:
                    break
            return cur

        with tempfile.TemporaryDirectory() as td:
            overlay_path = os.path.join(td, "overlay.pdf")
            base_reader = PdfReader(tpl_path)
            first_page = base_reader.pages[0]
            width = float(first_page.mediabox.width)
            height = float(first_page.mediabox.height)

            def to_pdf_coords(xy: tuple[float, float]) -> tuple[float, float]:
                # positions are stored in image pixels; preview_scale expresses pixels per PDF point
                px, py = float(xy[0]), float(xy[1])
                return (px / preview_scale, py / preview_scale)

            def norm_positions_dict(d: Dict[str, Any]) -> Dict[str, tuple[float, float]]:
                out = {}
                for k, v in d.items():
                    try:
                        out[k] = to_pdf_coords(v)
                    except Exception:
                        out[k] = v
                return out

            positions_pdf = norm_positions_dict(positions)
            header_pdf = norm_positions_dict(header_positions)
            footer_pdf = norm_positions_dict(footer_positions)

            images_cfg = mapping.get("_images", {}) or {}
            image_previews = mapping.get("_image_previews", {}) or {}
            signatures_cfg = mapping.get("_signatures", {}) or {}

            arr_path = None
            arr_def = None
            if repeat_rows_cfg:
                arr_path, arr_def = next(iter(repeat_rows_cfg.items()))
            items: List[Any] = get_path(original_data, arr_path) if arr_path else []
            if items is None or not isinstance(items, list):
                items = []

            rows_per_page = int(arr_def.get("rowsPerPage", 0)) if arr_def else 0
            start_y = float(arr_def.get("startY", 700)) if arr_def else 700.0
            delta_y = float(arr_def.get("deltaY", 24)) if arr_def else 24.0
            end_y = arr_def.get("endY") if arr_def else None
            if not rows_per_page and end_y is not None:
                try:
                    end_y = float(end_y)
                    if delta_y > 0:
                        rows_per_page = max(1, int((start_y - end_y) / delta_y) + 1)
                except Exception:
                    rows_per_page = 0
            if not rows_per_page:
                rows_per_page = max(1, int((start_y) / max(delta_y, 1)))

            total_pages = max(1, math.ceil(len(items) / rows_per_page))

            c = canvas.Canvas(overlay_path, pagesize=(width, height))

            def draw_header_footer(page_index: int):
                for key, (x, y) in header_pdf.items():
                    val = context.get(key)
                    if key == "_page_number":
                        val = str(page_index + 1)
                    if key == "_page_count":
                        val = str(total_pages)
                    draw_text(key, x, y, val)
                for key, (x, y) in footer_pdf.items():
                    val = context.get(key)
                    if key == "_page_number":
                        val = str(page_index + 1)
                    if key == "_page_count":
                        val = str(total_pages)
                    draw_text(key, x, y, val)

            def draw_fixed_positions():
                for key, value in context.items():
                    if key.startswith("_"):
                        continue
                    if arr_path and key.startswith(arr_path + "."):
                        continue
                    pos = positions_pdf.get(key)
                    if pos:
                        x, y = pos
                        draw_text(key, x, y, value)

            def decode_data_url(data_url_or_b64: str | bytes | None) -> bytes | None:
                if not data_url_or_b64:
                    return None
                if isinstance(data_url_or_b64, bytes):
                    return data_url_or_b64
                s = str(data_url_or_b64)
                try:
                    if s.startswith("data:") and ";base64," in s:
                        b64 = s.split(",",1)[1]
                        import base64
                        return base64.b64decode(b64)
                    # raw base64
                    import base64
                    return base64.b64decode(s)
                except Exception:
                    return None

            def draw_images():
                for key, meta in images_cfg.items():
                    # choose data: prefer context value, else preview
                    data_bytes = None
                    ctx_val = context.get(key)
                    
                    if ctx_val is not None:
                        # Check if it's a URL
                        if isinstance(ctx_val, str) and (ctx_val.startswith('http://') or ctx_val.startswith('https://')):
                            try:
                                import requests
                                print(f"📥 Descargando imagen desde: {ctx_val}")
                                response = requests.get(ctx_val, timeout=10, headers={'User-Agent': 'Mozilla/5.0'})
                                print(f"📊 Status: {response.status_code}, Content-Type: {response.headers.get('content-type', 'N/A')}")
                                
                                if response.status_code == 200:
                                    content_type = response.headers.get('content-type', '').lower()
                                    
                                    # Handle SVG files - convert to PNG or skip
                                    if 'svg' in content_type:
                                        # For SVG, try to convert or use a fallback
                                        try:
                                            # Try to convert SVG to PNG using cairosvg if available
                                            import cairosvg
                                            png_data = cairosvg.svg2png(bytestring=response.content)
                                            data_bytes = png_data
                                            print(f"✅ SVG convertido a PNG para {key}: {len(png_data)} bytes")
                                        except ImportError:
                                            # If cairosvg not available, skip SVG files
                                            print(f"⚠️ SVG no soportado para {key}, saltando")
                                            continue
                                    else:
                                        # For other image types, use as-is
                                        data_bytes = response.content
                                        print(f"✅ Imagen descargada para {key}: {len(data_bytes)} bytes")
                                else:
                                    print(f"❌ Error HTTP: {response.status_code}")
                                    continue
                            except Exception as e:
                                print(f"❌ Error descargando imagen para {key}: {e}")
                                continue
                        else:
                            # Try as data URL
                            data_bytes = decode_data_url(ctx_val)
                    
                    if not data_bytes:
                        data_bytes = decode_data_url(image_previews.get(key))
                    
                    if not data_bytes:
                        continue
                    
                    try:
                        img_reader = ImageReader(io.BytesIO(data_bytes))
                        print(f"✅ Imagen cargada para {key}: {img_reader.getSize()}")
                    except Exception as e:
                        print(f"❌ Error cargando imagen para {key}: {e}")
                        continue
                    
                    # convert px to pt and apply offset
                    x_pt = float(meta.get("x", 0.0)) / preview_scale + offset_x
                    y_pt = float(meta.get("y", 0.0)) / preview_scale + offset_y
                    w_pt = float(meta.get("width", 100.0)) / preview_scale
                    h_pt = float(meta.get("height", 100.0)) / preview_scale
                    # y_pt is the top-left anchor in our coordinate? We defined (x,y) as bottom-left of text; for images, assume (x,y) is bottom-left
                    # drawImage expects lower-left
                    try:
                        c.drawImage(img_reader, x_pt, y_pt - h_pt + h_pt, width=w_pt, height=h_pt, preserveAspectRatio=True, mask='auto')
                        print(f"✅ Imagen dibujada para {key} en ({x_pt:.1f}, {y_pt:.1f}) tamaño {w_pt:.1f}x{h_pt:.1f}")
                    except Exception as e:
                        print(f"❌ Error dibujando imagen para {key}: {e}")
                        continue

            def draw_signatures():
                for key, meta in signatures_cfg.items():
                    # convert px to pt and apply offset
                    x_pt = float(meta.get("x", 0.0)) / preview_scale + offset_x
                    y_pt = float(meta.get("y", 0.0)) / preview_scale + offset_y
                    w_pt = float(meta.get("width", 200.0)) / preview_scale
                    h_pt = float(meta.get("height", 300.0)) / preview_scale
                    
                    try:
                        # Draw signature rectangle
                        c.setLineWidth(1.0)
                        c.setStrokeColor(HexColor("#000000"))
                        c.rect(x_pt, y_pt - h_pt, w_pt, h_pt)
                        print(f"✅ Firma dibujada para {key} en ({x_pt:.1f}, {y_pt:.1f}) tamaño {w_pt:.1f}x{h_pt:.1f}")
                    except Exception as e:
                        print(f"❌ Error dibujando firma para {key}: {e}")
                        continue

            for page_idx in range(total_pages):
                draw_header_footer(page_idx)
                draw_fixed_positions()
                draw_images()
                draw_signatures()
                debug_lines = []
                debug_lines.append(f"scale={preview_scale}, offset=({offset_x:.1f},{offset_y:.1f}), page= {page_idx+1}/{total_pages}, size=({width:.1f},{height:.1f})")
                # Fixed fields debug
                for key, (x_pt, y_pt) in positions_pdf.items():
                    if arr_path and key.startswith(arr_path + "."):
                        continue
                    final_x = x_pt + offset_x
                    final_y = y_pt + offset_y
                    x_px, y_px = positions.get(key, (None, None))
                    if x_px is None:
                        continue
                    debug_lines.append(f"{key}: px=({x_px:.1f},{y_px:.1f}) -> pt=({x_pt:.1f},{y_pt:.1f}) final=({final_x:.1f},{final_y:.1f})")
                    # draw crosshair at final position for visual check
                    draw_cross(x_pt, y_pt)
                # Repeat rows debug (only if array defined)
                if arr_path:
                    start_idx = page_idx * rows_per_page
                    end_idx = min(len(items), start_idx + rows_per_page)
                    for idx in range(start_idx, end_idx):
                        y_item = start_y - (idx - start_idx) * delta_y
                        prefix = arr_path + "."
                        for key, (x_pt, y_pt_base) in positions_pdf.items():
                            if key.startswith(prefix):
                                final_x = x_pt + offset_x
                                final_y = y_item + offset_y
                                x_px, y_px = positions.get(key, (None, None))
                                if x_px is None:
                                    continue
                                debug_lines.append(f"{key}[{idx}]: px=({x_px:.1f},{y_px:.1f}) -> base_pt=({x_pt:.1f},{y_pt_base if isinstance(y_pt_base,(int,float)) else 0:.1f}) y_item={y_item:.1f} final=({final_x:.1f},{final_y:.1f})")
                                draw_cross(x_pt, y_item)
                # Draw debug lines in gray at top-left
                try:
                    c.setFont("Helvetica", 6)
                except Exception:
                    pass
                try:
                    c.setFillColor(HexColor("#666666"))
                except Exception:
                    pass
                y_cursor = height - 10
                for line in debug_lines[:100]:
                    c.drawString(10, y_cursor, line)
                    y_cursor -= 8
                # Draw repeated rows content
                if arr_path:
                    start_idx = page_idx * rows_per_page
                    end_idx = min(len(items), start_idx + rows_per_page)
                    for idx in range(start_idx, end_idx):
                        item = items[idx]
                        y_item = start_y - (idx - start_idx) * delta_y
                        prefix = arr_path + "."
                        for key, (x, _) in positions_pdf.items():
                            if key.startswith(prefix):
                                sub_key = key[len(prefix):]
                                val = get_path(item, sub_key)
                                draw_text(key, x, y_item, val)
                if page_idx < total_pages - 1:
                    c.showPage()
            c.save()

            overlay_reader = PdfReader(overlay_path)
            writer = PdfWriter()
            base_page_count = len(base_reader.pages)
            for i in range(total_pages):
                base_page = base_reader.pages[min(i, base_page_count - 1)]
                merged = base_page
                if i < len(overlay_reader.pages):
                    merged.merge_page(overlay_reader.pages[i])
                writer.add_page(merged)

            out_path = os.path.join(td, "out.pdf")
            with open(out_path, "wb") as f:
                writer.write(f)
            with open(out_path, "rb") as f:
                return f.read()

    async def convert_pdf_to_image(self, pdf_bytes: bytes) -> bytes:
        """
        Convierte un PDF a imagen PNG optimizada.
        
        Args:
            pdf_bytes: Bytes del PDF a convertir
            
        Returns:
            Bytes de la imagen PNG
        """
        try:
            # Usar un executor para la conversión síncrona
            loop = asyncio.get_event_loop()
            with ThreadPoolExecutor() as executor:
                image_bytes = await loop.run_in_executor(
                    executor, 
                    self._convert_pdf_to_image_sync, 
                    pdf_bytes
                )
            return image_bytes
        except Exception as e:
            raise RuntimeError(f"Error convirtiendo PDF a imagen: {e}")

    def _convert_pdf_to_image_sync(self, pdf_bytes: bytes) -> bytes:
        """
        Conversión síncrona de PDF a imagen PNG.
        
        Args:
            pdf_bytes: Bytes del PDF a convertir
            
        Returns:
            Bytes de la imagen PNG
        """
        print(f"🔄 Iniciando conversión PDF a imagen (tamaño PDF: {len(pdf_bytes)} bytes)")
        
        try:
            print("🔄 Intentando conversión con LibreOffice...")
            result = self._convert_pdf_to_image_libreoffice(pdf_bytes)
            print(f"✅ LibreOffice conversion successful (tamaño imagen: {len(result)} bytes)")
            return result
        except Exception as e:
            print(f"⚠️  LibreOffice conversion failed: {e}")
            try:
                print("🔄 Intentando conversión con pdf2image...")
                result = self._convert_pdf_to_image_pil(pdf_bytes)
                print(f"✅ pdf2image conversion successful (tamaño imagen: {len(result)} bytes)")
                return result
            except Exception as e2:
                print(f"⚠️  pdf2image conversion failed: {e2}")
                print("🔄 Usando método de fallback...")
                # Último recurso: imagen de error
                from PIL import Image, ImageDraw
                import io
                
                img = Image.new('RGB', (400, 200), color='lightgray')
                draw = ImageDraw.Draw(img)
                draw.text((50, 80), "Error: No se pudo convertir PDF", fill='red')
                
                img_io = io.BytesIO()
                img.save(img_io, format='PNG')
                img_io.seek(0)
                result = img_io.getvalue()
                print(f"⚠️  Fallback method used (tamaño imagen: {len(result)} bytes)")
                return result

    def _convert_pdf_to_image_libreoffice(self, pdf_bytes: bytes) -> bytes:
        """
        Convierte PDF a imagen usando LibreOffice (conversión real).
        
        Args:
            pdf_bytes: Bytes del PDF a convertir
            
        Returns:
            Bytes de la imagen PNG
        """
        import tempfile
        import os
        import subprocess
        
        print("🔄 Intentando conversión con LibreOffice...")
        
        # Verificar si LibreOffice está disponible
        try:
            result = subprocess.run(['soffice', '--version'], capture_output=True, text=True)
            if result.returncode != 0:
                print(f"❌ LibreOffice no está disponible: {result.stderr}")
                raise RuntimeError("LibreOffice no está instalado o no está en el PATH")
            print(f"✅ LibreOffice encontrado: {result.stdout.strip()}")
        except FileNotFoundError:
            print("❌ LibreOffice no está en el PATH")
            raise RuntimeError("LibreOffice no está instalado o no está en el PATH")
        
        with tempfile.TemporaryDirectory() as temp_dir:
            # Guardar PDF temporal
            pdf_path = os.path.join(temp_dir, "input.pdf")
            with open(pdf_path, "wb") as f:
                f.write(pdf_bytes)
            print(f"📄 PDF guardado en: {pdf_path}")
            
            # Convertir a imagen usando LibreOffice directamente
            image_path = os.path.join(temp_dir, "output.png")
            cmd = [
                'soffice', 
                '--headless', 
                '--convert-to', 'png', 
                '--outdir', temp_dir, 
                pdf_path
            ]
            
            print(f"🔄 Ejecutando comando: {' '.join(cmd)}")
            result = subprocess.run(cmd, capture_output=True, text=True)
            
            if result.returncode == 0:
                # LibreOffice genera el archivo con un nombre diferente
                expected_image_path = os.path.join(temp_dir, "input.png")
                if os.path.exists(expected_image_path):
                    print(f"✅ Imagen generada: {expected_image_path}")
                    with open(expected_image_path, "rb") as f:
                        return f.read()
                else:
                    print(f"❌ Archivo de imagen no encontrado en: {expected_image_path}")
                    print(f"📁 Archivos en el directorio: {os.listdir(temp_dir)}")
                    raise RuntimeError("LibreOffice no generó el archivo de imagen")
            else:
                print(f"❌ Error en LibreOffice: {result.stderr}")
                raise RuntimeError(f"LibreOffice falló: {result.stderr}")

    def _convert_pdf_to_image_pil(self, pdf_bytes: bytes) -> bytes:
        """
        Convierte PDF a imagen usando pypdfium2 (sin dependencias externas).
        
        Args:
            pdf_bytes: Bytes del PDF a convertir
            
        Returns:
            Bytes de la imagen PNG
        """
        print("🔄 Intentando conversión con pypdfium2...")
        
        try:
            import pypdfium2 as pdfium
            from PIL import Image
            import io
            
            print(f"📄 Convirtiendo PDF de {len(pdf_bytes)} bytes...")
            
            # Cargar PDF con pypdfium2
            pdf = pdfium.PdfDocument(pdf_bytes)
            print(f"✅ PDF cargado con {len(pdf)} páginas")
            
            if len(pdf) > 0:
                # Renderizar primera página
                page = pdf[0]
                bitmap = page.render(
                    scale=2.0,  # 2x zoom para mejor calidad
                    rotation=0,
                    crop=(0, 0, 0, 0)
                )
                
                # Convertir a PIL Image
                pil_image = bitmap.to_pil()
                print(f"🖼️  Tamaño de imagen: {pil_image.size}, Modo: {pil_image.mode}")
                
                # Convertir a RGB si es necesario
                if pil_image.mode != 'RGB':
                    pil_image = pil_image.convert('RGB')
                    print("🔄 Convertida a RGB")
                
                # Optimizar y guardar
                img_io = io.BytesIO()
                pil_image.save(img_io, format='PNG', optimize=True)
                img_io.seek(0)
                result = img_io.getvalue()
                print(f"✅ Imagen PNG generada: {len(result)} bytes")
                return result
            else:
                print("❌ PDF no tiene páginas")
                raise RuntimeError("PDF no tiene páginas")
                
        except ImportError as e:
            print(f"❌ pypdfium2 no está disponible: {e}")
            return self._convert_pdf_to_image_fallback(pdf_bytes)
        except Exception as e:
            print(f"❌ Error en pypdfium2: {e}")
            import traceback
            print(f"📋 Traceback: {traceback.format_exc()}")
            return self._convert_pdf_to_image_fallback(pdf_bytes)

    def _convert_pdf_to_image_fallback(self, pdf_bytes: bytes) -> bytes:
        """
        Método de respaldo para convertir PDF a imagen usando PIL.
        
        Args:
            pdf_bytes: Bytes del PDF a convertir
            
        Returns:
            Bytes de la imagen PNG
        """
        try:
            from PIL import Image, ImageDraw, ImageFont
            import io
            
            # Crear una imagen que simule una página de documento
            img = Image.new('RGB', (800, 1000), color='white')
            draw = ImageDraw.Draw(img)
            
            # Intentar extraer texto del PDF
            pdf_reader = PdfReader(io.BytesIO(pdf_bytes))
            text_content = ""
            
            for page in pdf_reader.pages:
                text_content += page.extract_text() or ""
            
            # Dibujar un encabezado
            try:
                font_large = ImageFont.load_default()
                font_small = ImageFont.load_default()
            except:
                font_large = None
                font_small = None
            
            # Dibujar título
            draw.text((50, 30), "DOCUMENTO GENERADO", fill='darkblue', font=font_large)
            
            # Dibujar línea separadora
            draw.line([(50, 60), (750, 60)], fill='darkblue', width=2)
            
            # Dibujar el contenido del PDF (limitado a 800 caracteres)
            if text_content:
                # Dividir el texto en líneas
                lines = []
                current_line = ""
                words = text_content[:800].split()
                
                for word in words:
                    if len(current_line + " " + word) < 80:
                        current_line += " " + word if current_line else word
                    else:
                        if current_line:
                            lines.append(current_line)
                        current_line = word
                
                if current_line:
                    lines.append(current_line)
                
                # Dibujar las líneas
                y_position = 80
                for line in lines[:30]:  # Máximo 30 líneas
                    draw.text((50, y_position), line, fill='black', font=font_small)
                    y_position += 20
                
                if len(lines) > 30:
                    draw.text((50, y_position), "...", fill='gray', font=font_small)
            else:
                draw.text((50, 80), "Documento generado correctamente", fill='black', font=font_small)
                draw.text((50, 100), "El contenido del PDF se ha procesado exitosamente", fill='gray', font=font_small)
            
            # Dibujar pie de página
            draw.line([(50, 950), (750, 950)], fill='darkblue', width=1)
            draw.text((50, 970), "GenDoc Service - Conversión PDF a Imagen", fill='gray', font=font_small)
            
            # Convertir a bytes optimizados
            img_io = io.BytesIO()
            img.save(img_io, format='PNG', optimize=True, quality=85)
            img_io.seek(0)
            return img_io.getvalue()
            
        except Exception as e:
            # Último recurso: imagen de error
            from PIL import Image, ImageDraw
            import io
            
            img = Image.new('RGB', (400, 200), color='lightgray')
            draw = ImageDraw.Draw(img)
            draw.text((50, 80), "Error: No se pudo convertir PDF", fill='red')
            
            img_io = io.BytesIO()
            img.save(img_io, format='PNG')
            img_io.seek(0)
            return img_io.getvalue()